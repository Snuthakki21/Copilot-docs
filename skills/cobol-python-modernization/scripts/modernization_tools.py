from __future__ import annotations

import argparse
import ast
import json
import re
from pathlib import Path
from typing import Iterable

_HEADER_ALIASES = {
    "job": "job", "jobname": "job", "job_name": "job", "job name": "job", "jobs": "job",
    "step": "step", "steps": "step", "stepname": "step", "step_name": "step", "step name": "step",
    "program": "program", "programs": "program", "pgm": "program",
    "input": "inputs", "inputs": "inputs", "inputfile": "inputs", "inputfiles": "inputs", "input files": "inputs",
    "output": "outputs", "outputs": "outputs", "outputfile": "outputs", "outputfiles": "outputs", "output files": "outputs",
}

_REQUIRED_FILES = (
    "lineage.md", "job.py", "run.jil", "oracle.sql", "sqlite.sql",
    "validation_report.md", "traceability.json", "tests/test_job.py",
)
_DOC_SECTIONS = ("Purpose:", "Inputs:", "Outputs:", "Side effects:", "Failure behavior:", "Rule IDs:")


def _header_key(value: object) -> str:
    text = "" if value is None else str(value).strip().lower()
    return re.sub(r"[^a-z0-9]+", " ", text).strip()


def normalize_headers(headers: Iterable[object]) -> list[str]:
    result: list[str] = []
    for header in headers:
        key = _header_key(header)
        compact = key.replace(" ", "")
        result.append(_HEADER_ALIASES.get(key, _HEADER_ALIASES.get(compact, key.replace(" ", "_"))))
    return result


def safe_job_name(value: object) -> str:
    name = str(value).strip()
    if not name or name in {".", ".."} or "/" in name or "\\" in name or ".." in name:
        raise ValueError(f"unsafe job name: {value!r}")
    if not re.fullmatch(r"[A-Za-z0-9._-]+", name):
        raise ValueError(f"unsupported job-name characters: {value!r}")
    return name


def workbook_rows(path: Path, sheet: str | None = None) -> list[dict[str, object]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError("openpyxl is required to read .xlsx inputs") from exc

    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb[sheet] if sheet else wb[wb.sheetnames[0]]
        rows = ws.iter_rows(values_only=True)
        try:
            headers = normalize_headers(next(rows))
        except StopIteration:
            return []
        required = {"job", "step", "program", "inputs", "outputs"}
        missing = sorted(required - set(headers))
        if missing:
            raise ValueError(f"missing required Excel columns: {', '.join(missing)}")
        records: list[dict[str, object]] = []
        for row in rows:
            if not any(value not in (None, "") for value in row):
                continue
            record = dict(zip(headers, row))
            record["job"] = safe_job_name(record["job"])
            records.append(record)
        return records
    finally:
        wb.close()


def _python_documentation_errors(path: Path) -> list[str]:
    if not path.is_file():
        return []
    try:
        tree = ast.parse(path.read_text(encoding="utf-8"))
    except (OSError, SyntaxError) as exc:
        return [f"job.py cannot be parsed for documentation validation: {exc}"]

    errors: list[str] = []
    if not ast.get_docstring(tree):
        errors.append("job.py has no module-level plain-English docstring")
    for node in ast.walk(tree):
        if isinstance(node, (ast.FunctionDef, ast.AsyncFunctionDef)):
            doc = ast.get_docstring(node) or ""
            missing = [section for section in _DOC_SECTIONS if section not in doc]
            if missing:
                errors.append(f"function {node.name} documentation missing: {', '.join(missing)}")
    return errors


def validate_job_artifacts(job_dir: Path) -> dict[str, object]:
    errors: list[str] = []
    for rel in _REQUIRED_FILES:
        if not (job_dir / rel).is_file():
            errors.append(f"missing required artifact: {rel}")

    errors.extend(_python_documentation_errors(job_dir / "job.py"))

    manifest_path = job_dir / "traceability.json"
    manifest: dict[str, object] = {}
    if manifest_path.is_file():
        try:
            manifest = json.loads(manifest_path.read_text(encoding="utf-8"))
        except (json.JSONDecodeError, OSError) as exc:
            errors.append(f"invalid traceability.json: {exc}")

    for key, label in (("source_commit", "source commit"), ("workbook_sha256", "workbook SHA-256")):
        value = manifest.get(key) if isinstance(manifest, dict) else None
        if not isinstance(value, str) or not value.strip():
            errors.append(f"traceability.json has no {label} provenance")

    rules = manifest.get("rules", []) if isinstance(manifest, dict) else []
    if not isinstance(rules, list) or not rules:
        errors.append("traceability.json must contain at least one discovered source rule")
    else:
        for idx, rule in enumerate(rules, start=1):
            if not isinstance(rule, dict):
                errors.append(f"rule #{idx} is not an object")
                continue
            rule_id = rule.get("id") or f"#{idx}"
            if not rule.get("source"):
                errors.append(f"rule {rule_id} has no source trace")
            if not rule.get("python"):
                errors.append(f"rule {rule_id} has no Python trace")
            if not rule.get("tests"):
                errors.append(f"rule {rule_id} has no validation test trace")

    coverage = manifest.get("coverage_percent", 0) if isinstance(manifest, dict) else 0
    if not isinstance(coverage, (int, float)) or coverage < 80:
        errors.append(f"Python coverage is below 80%: {coverage!r}")

    for key, label in (("security_findings", "security findings"), ("defects", "defects"), ("unresolved", "unresolved source behavior")):
        value = manifest.get(key, []) if isinstance(manifest, dict) else []
        if value:
            errors.append(f"non-zero {label}: {len(value) if isinstance(value, list) else value}")

    return {"passed": not errors, "score": 100 if not errors else 0, "errors": errors}


def main() -> int:
    parser = argparse.ArgumentParser(description="Validate one generated modernization job folder.")
    parser.add_argument("job_dir", type=Path)
    args = parser.parse_args()
    result = validate_job_artifacts(args.job_dir)
    print(json.dumps(result, indent=2))
    return 0 if result["passed"] else 1


if __name__ == "__main__":
    raise SystemExit(main())
