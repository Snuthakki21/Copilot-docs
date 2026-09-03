import json
import tempfile
import unittest
from pathlib import Path

from scripts.modernization_tools import normalize_headers, safe_job_name, validate_job_artifacts, workbook_rows

DOC_JOB = '''"""Plain-English job summary."""\n\ndef run_job():\n    """Purpose: Run the job.\n    Inputs: None.\n    Outputs: Exit result.\n    Side effects: Writes validated outputs.\n    Failure behavior: Raises on invalid processing.\n    Rule IDs: R1.\n    """\n    return 0\n'''

VALIDATION_REPORT = '''# Executive Validation Summary\nPASS\n\n# Rule Parity Matrix\nRule ID: R1\nSource code evidence: JCL:STEP01\nSource behavior in plain English: The source runs rule R1.\nPython code evidence: job.py:3\nPython behavior in plain English: Python runs the same rule.\nParity result: PASS\nDiscrepancy: None\nRoot cause: N/A\nRequired remediation: None\n\n# Preserved Legacy Quirks\nNone.\n\n# File / Database / Scheduler Reconciliation\nPASS\n\n# Test / Coverage / Security Results\nPASS\n\n# Final Verdict\n100%\n'''


def make_complete(root: Path, coverage=80):
    for name in ["lineage.md", "run.jil", "oracle.sql", "sqlite.sql"]:
        (root / name).write_text("placeholder")
    (root / "validation_report.md").write_text(VALIDATION_REPORT)
    (root / "job.py").write_text(DOC_JOB)
    (root / "tests").mkdir()
    (root / "tests" / "test_job.py").write_text("def test_x(): assert True")
    manifest = {
        "source_commit": "abc123",
        "workbook_sha256": "0" * 64,
        "rules": [{"id": "R1", "source": "JCL:STEP01", "python": ["job.py:3"], "tests": ["tests/test_job.py::test_r1"]}],
        "coverage_percent": coverage,
        "security_findings": [], "defects": [], "unresolved": [], "unsupported_claims": [],
    }
    (root / "traceability.json").write_text(json.dumps(manifest))
    return manifest


class ToolingTests(unittest.TestCase):
    def test_normalize_headers_accepts_common_excel_variants(self):
        self.assertEqual(normalize_headers(["Job Name", "STEP", "Program", "Input Files", "Outputs"]), ["job", "step", "program", "inputs", "outputs"])

    def test_safe_job_name_rejects_path_traversal_and_normalizes(self):
        self.assertEqual(safe_job_name(" AR.DAILY-01 "), "AR.DAILY-01")
        with self.assertRaises(ValueError):
            safe_job_name("../escape")

    def test_validator_fails_closed_when_required_artifact_missing(self):
        with tempfile.TemporaryDirectory() as td:
            root = Path(td)
            (root / "traceability.json").write_text(json.dumps({"rules": []}))
            result = validate_job_artifacts(root)
            self.assertFalse(result["passed"])
            self.assertIn("lineage.md", "\n".join(result["errors"]))

    def test_validator_requires_rule_test_trace(self):
        with tempfile.TemporaryDirectory() as td:
            root = Path(td)
            manifest = make_complete(root, 90)
            manifest["rules"][0]["tests"] = []
            (root / "traceability.json").write_text(json.dumps(manifest))
            self.assertTrue(any("R1" in e for e in validate_job_artifacts(root)["errors"]))

    def test_validator_accepts_complete_manifest(self):
        with tempfile.TemporaryDirectory() as td:
            root = Path(td)
            make_complete(root)
            result = validate_job_artifacts(root)
            self.assertTrue(result["passed"], result)

    def test_validator_rejects_missing_provenance(self):
        with tempfile.TemporaryDirectory() as td:
            root = Path(td)
            manifest = make_complete(root)
            del manifest["source_commit"]
            (root / "traceability.json").write_text(json.dumps(manifest))
            self.assertTrue(any("source commit" in e for e in validate_job_artifacts(root)["errors"]))

    def test_validator_rejects_undocumented_function(self):
        with tempfile.TemporaryDirectory() as td:
            root = Path(td)
            make_complete(root)
            (root / "job.py").write_text('"""Job."""\ndef run_job():\n    return 0\n')
            text = "\n".join(validate_job_artifacts(root)["errors"])
            self.assertIn("function run_job documentation missing", text)

    def test_validator_rejects_missing_module_docstring(self):
        with tempfile.TemporaryDirectory() as td:
            root = Path(td)
            make_complete(root)
            (root / "job.py").write_text('def run_job():\n    """Purpose: x\nInputs: x\nOutputs: x\nSide effects: x\nFailure behavior: x\nRule IDs: R1\n"""\n    return 0\n')
            self.assertTrue(any("module-level" in e for e in validate_job_artifacts(root)["errors"]))

    def test_validator_rejects_incomplete_validation_report(self):
        with tempfile.TemporaryDirectory() as td:
            root = Path(td)
            make_complete(root)
            (root / "validation_report.md").write_text("PASS")
            text = "\n".join(validate_job_artifacts(root)["errors"])
            self.assertIn("Rule Parity Matrix", text)
            self.assertIn("Source behavior in plain English", text)

    def test_validator_requires_every_rule_in_validation_report(self):
        with tempfile.TemporaryDirectory() as td:
            root = Path(td)
            manifest = make_complete(root)
            manifest["rules"].append({"id": "R2", "source": "COBOL:P2", "python": ["job.py:3"], "tests": ["tests/test_job.py::test_r2"]})
            (root / "traceability.json").write_text(json.dumps(manifest))
            self.assertTrue(any("rule R2" in e for e in validate_job_artifacts(root)["errors"]))

    def test_validator_rejects_unsupported_validation_claims(self):
        with tempfile.TemporaryDirectory() as td:
            root = Path(td)
            manifest = make_complete(root)
            manifest["unsupported_claims"] = ["R1 root cause asserted without evidence"]
            (root / "traceability.json").write_text(json.dumps(manifest))
            self.assertTrue(any("unsupported validation claims" in e for e in validate_job_artifacts(root)["errors"]))

    def test_workbook_rows_reads_expected_columns_and_skips_blank_rows(self):
        from openpyxl import Workbook
        with tempfile.TemporaryDirectory() as td:
            path = Path(td) / "jobs.xlsx"
            wb = Workbook()
            ws = wb.active
            ws.append(["Jobs", "Steps", "Programs", "Input Files", "Outputs"])
            ws.append(["JOB1", "STEP1", "PGM1", "IN1", "OUT1"])
            ws.append([None] * 5)
            wb.save(path)
            rows = workbook_rows(path)
            self.assertEqual(len(rows), 1)
            self.assertEqual(rows[0]["job"], "JOB1")

    def test_workbook_rows_rejects_missing_columns(self):
        from openpyxl import Workbook
        with tempfile.TemporaryDirectory() as td:
            path = Path(td) / "jobs.xlsx"
            wb = Workbook()
            ws = wb.active
            ws.append(["Job", "Step", "Program"])
            wb.save(path)
            with self.assertRaisesRegex(ValueError, "missing required Excel columns"):
                workbook_rows(path)

    def test_workbook_rows_empty_sheet_returns_empty(self):
        from openpyxl import Workbook
        with tempfile.TemporaryDirectory() as td:
            path = Path(td) / "jobs.xlsx"
            wb = Workbook()
            wb.save(path)
            self.assertEqual(workbook_rows(path), [])

    def test_validator_rejects_invalid_json_low_coverage_and_findings(self):
        with tempfile.TemporaryDirectory() as td:
            root = Path(td)
            make_complete(root)
            (root / "traceability.json").write_text("not json")
            self.assertFalse(validate_job_artifacts(root)["passed"])
        with tempfile.TemporaryDirectory() as td:
            root = Path(td)
            manifest = make_complete(root, 79.9)
            manifest["security_findings"] = ["x"]
            manifest["defects"] = ["y"]
            manifest["unresolved"] = ["z"]
            (root / "traceability.json").write_text(json.dumps(manifest))
            text = "\n".join(validate_job_artifacts(root)["errors"])
            self.assertIn("below 80%", text)
            self.assertIn("security findings", text)
            self.assertIn("defects", text)
            self.assertIn("unresolved source behavior", text)


if __name__ == "__main__":
    unittest.main()
